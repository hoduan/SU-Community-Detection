'''
   This code is designed to read community information from .xlsx and then convert it into a dictionary structure.
   Do not copy, disclose, or distribute without explicit written permission. 
   Author: 			Hongyu Duan <duanhongyu2010@gmail.com>
   
 
   usage: In Linux command line using the following command:
	   python helper.py [Input File Name1] [Kcore/Rcore value] / there should be a sheet named "Data" in the input .xlsx file
'''

import sys
from sklearn.metrics.cluster import normalized_mutual_info_score
import collections
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def main():
	infile = sys.argv[1]
	#outfile = str(sys.argv[2]) + "_" + sys.argv[3]
	temp={}
	k = int(sys.argv[2])
	
	
	#open the input file to read data
	wb = openpyxl.load_workbook(infile)
	sheet = wb.get_sheet_by_name('Data')
	
	row_count = sheet.max_row	
	#for row in sheet.iter_rows():
		#temphash[sheet["B"+row].value] = 
		#	print cell.value
	#	print row
	
	for rows in range(2,row_count+1):
		#temp[sheet.cell('A'+str(row))] = sheet.cell('B' + str(row))
		temp[int(sheet.cell(row = rows, column=1).value)] = int(sheet.cell(row = rows, column = k+1).value)	
	print temp

if __name__ == "__main__":main()
