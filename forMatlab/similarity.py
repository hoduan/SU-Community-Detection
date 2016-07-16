'''
   This code is designed and written by Hongyu Duan.
   Do not copy, disclose, or distribute without explicit written permission. 
   Author: 			Hongyu Duan <duanhongyu2010@gmail.com>
   
   Instruction:			This program is a python package which is designed to compute the normalized similarity between two community partitions
				The Input File should hold data in the structure of dictionary with community node as the key, and its corresponding community lable as the value. 
 
   usage: In Linux command line using the following command:
	   python similarity.py [Input File Name1] [Input File Name2] [Dataset] [Kcore_Value]
'''

import sys
from sklearn.metrics.cluster import normalized_mutual_info_score
import collections
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

def main():
	file1 = sys.argv[1]
	file2 = sys.argv[2]
	
	c_true = {}
	c_pred = {}
	#read data from file
	with open(file1) as fd1, open(file2) as fd2:
		c_true = eval(fd1.readline())
		c_pred = eval(fd2.readline())
	#order the data in dictionary data structure
	c_true_order = collections.OrderedDict(sorted(c_true.items()))
	c_pred_order = collections.OrderedDict(sorted(c_pred.items()))
	c_true_label = []
	c_pred_label = []
	print c_true_order	
	#make list with community label 
	for k, v in c_true_order.items():
		c_true_label.append(v)
	for k, v in c_pred_order.items():
		c_pred_label.append(v)
	
	
	simi =  normalized_mutual_info_score(c_true_label,c_pred_label)
	

	DATA_FILE = sys.argv[3].split("/")
	FILE_LOG_NAME = "LOG_File_"+(DATA_FILE[-1])+ ".xlsx"
	Kcore_Value = int(sys.argv[4])

	if(not os.path.exists(FILE_LOG_NAME)):
		wb = openpyxl.Workbook()	
		sheet = wb.active
		sheet.title = "Sheet1"
		sheet['A1'] = 'K/R Value'
		sheet['B1'] = 'NMI Similarity'
		sheet['A2'] = 'v=10%'
		sheet['A3'] = 'v=20%'
		sheet['A4'] = 'v=30%'
		sheet['A5'] = 'v=40%'
		sheet['A6'] = 'v=50%'
		sheet['A7'] = 'v=60%'
		sheet['A8'] = 'v=70%'
		sheet['A9'] = 'v=80%'
		sheet['A10'] = 'v=90%'
		sheet['A11'] = 'v=100%'
	else:
		wb = openpyxl.load_workbook(FILE_LOG_NAME)
        
	sheet = wb.get_sheet_by_name('Sheet1')
	sheet['B'+str(Kcore_Value + 1)] = simi
	wb.save(FILE_LOG_NAME)

if __name__ == "__main__":main()
